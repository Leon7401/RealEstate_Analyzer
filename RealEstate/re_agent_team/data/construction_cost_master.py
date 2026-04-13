"""context配下の原価資料からコストマスタを推定するユーティリティ。"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Optional
import logging


logger = logging.getLogger("ConstructionCostMaster")


@dataclass
class StructureCostProfile:
    cost_per_sqm: int
    overhead_rate: float
    fixed_cost: int = 0
    demolition_cost_per_sqm: int = 25_000


def _safe_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").replace("円", "").replace("%", "").strip())
    except Exception:
        return None


def _sheet_total_cost_yen(ws) -> Optional[int]:
    """「合計」行の数値を走査し、現実的な総工費を推定する。"""
    candidates = []
    for row in ws.iter_rows(min_row=1, max_row=800, values_only=True):
        row_text = " ".join(str(v) for v in row if isinstance(v, str))
        if "合計" not in row_text and "総計" not in row_text:
            continue
        nums = [v for v in row if isinstance(v, (int, float))]
        for n in nums:
            if 2_000_000 <= float(n) <= 200_000_000:
                candidates.append(float(n))
    if not candidates:
        return None
    return int(max(candidates))


def _sheet_reference_floor_area(ws) -> Optional[float]:
    """
    シート先頭付近の面積セル（実務シートの典型値）から参照延床を推定。
    28.6 のような戸当たり面積が入るため、6戸換算で最低延床を組む。
    """
    vals = []
    # 先頭数行に「延床相当」が置かれているケースを優先
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        for v in row:
            fv = _safe_float(v)
            if fv is None:
                continue
            if 80 <= fv <= 2_000:
                vals.append(fv)
    if vals:
        return max(vals)

    # 見つからない場合は戸当たり面積(20-60㎡)を抽出して6戸換算
    vals = []
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        for v in row:
            fv = _safe_float(v)
            if fv is None:
                continue
            if 18 <= fv <= 60:
                vals.append(fv)
    if not vals:
        return None
    base = sorted(vals)[len(vals) // 2]
    return max(base * 6, 120.0)


def _load_profiles_from_cost_analysis(xlsx_path: Path) -> Dict[str, StructureCostProfile]:
    profiles: Dict[str, StructureCostProfile] = {}
    try:
        from openpyxl import load_workbook
    except Exception:
        logger.warning("openpyxl未導入のため原価分析シートを読み込めません")
        return profiles

    if not xlsx_path.exists():
        return profiles

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet_map = {
        "木三共_原価分析": "木造",
        "長屋_原価分析": "木造",
        "Kフレーム_原価分析": "重量鉄骨",
        "重量鉄骨_原価分析": "重量鉄骨",
    }
    per_structure_costs: Dict[str, list[int]] = {"木造": [], "重量鉄骨": []}

    for sheet_name, structure in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        total_cost = _sheet_total_cost_yen(ws)
        floor_area = _sheet_reference_floor_area(ws)
        if not total_cost or not floor_area or floor_area <= 0:
            continue
        per_sqm = int(total_cost / floor_area)
        if 60_000 <= per_sqm <= 900_000:
            per_structure_costs[structure].append(per_sqm)

    for structure, vals in per_structure_costs.items():
        if not vals:
            continue
        vals.sort()
        median = vals[len(vals) // 2]
        profiles[structure] = StructureCostProfile(
            cost_per_sqm=median,
            overhead_rate=0.15,
        )

    return profiles


def _estimate_fixed_and_demo_cost(project_xlsx: Path) -> tuple[int, int]:
    """
    「工期・原価・発注先」シートから、固定コストと解体単価を推定。
    固定コスト: 概算費用(万円)が入る行を合計。
    解体単価: 解体関連行の(概算費用/数量)を㎡単価化。
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        return 0, 25_000
    if not project_xlsx.exists():
        return 0, 25_000

    wb = load_workbook(project_xlsx, data_only=True, read_only=True)
    if "費用｜木造｜9戸｜25㎡" not in wb.sheetnames:
        return 0, 25_000
    ws = wb["費用｜木造｜9戸｜25㎡"]

    fixed_total_yen = 0.0
    demo_unit_candidates = []
    for row in ws.iter_rows(min_row=1, max_row=500, values_only=True):
        text_cells = [str(v) for v in row if isinstance(v, str)]
        line_text = " ".join(text_cells)
        nums = [float(v) for v in row if isinstance(v, (int, float))]
        if not nums:
            continue

        # 概算費用(万円)を末尾側から優先採用
        man = None
        for n in reversed(nums):
            if 0 < n < 100_000:
                man = n
                break
        if man is not None:
            fixed_total_yen += man * 10_000

        if "解体" in line_text and len(nums) >= 2:
            qty = nums[-2]
            cost_man = nums[-1] if nums[-1] < 100_000 else nums[-2]
            if qty and qty > 0 and cost_man > 0:
                unit = (cost_man * 10_000) / qty
                if 5_000 <= unit <= 120_000:
                    demo_unit_candidates.append(unit)

    demolition = int(sum(demo_unit_candidates) / len(demo_unit_candidates)) if demo_unit_candidates else 25_000
    # 読み取り誤差を避けるため、固定費は常識レンジでクリップ
    fixed = int(max(0, min(fixed_total_yen, 8_000_000)))
    return fixed, demolition


def load_construction_cost_profiles(base_dir: Path) -> Dict[str, StructureCostProfile]:
    """
    context資料から構造別コストプロファイルを返す。
    見つからない場合は空dict（呼び出し側で既定値を使用）。
    """
    context_dir = base_dir.parent / "context"
    analysis_xlsx = context_dir / "参考｜原価分析.xlsx"
    project_xlsx = context_dir / "工期・原価・発注先.xlsx"

    profiles = _load_profiles_from_cost_analysis(analysis_xlsx)
    fixed_cost, demolition_cost = _estimate_fixed_and_demo_cost(project_xlsx)
    for profile in profiles.values():
        profile.fixed_cost = fixed_cost
        profile.demolition_cost_per_sqm = demolition_cost
        # 実務原価を反映した付帯費をやや厚めに補正
        profile.overhead_rate = max(profile.overhead_rate, 0.18)
    return profiles

